"""
Handys – Excel-Export
Erstellt eine formatierte Übersichtsdatei aller Diensthandys im DRK-Design.
"""
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import HANDYS_EXPORT_PATH

# ─── DRK Corporate-Design Farben ─────────────────────────────────────────────
DRK_ROT        = "CC0000"
DRK_DUNKELGRAU = "404040"
WEISS          = "FFFFFF"
HELLGRAU       = "F2F2F2"
MITTELGRAU     = "D9D9D9"

ZUSTAND_FARBEN = {
    "Aktiv":         {"bg": "C6EFCE", "fg": "276221"},
    "Defekt":        {"bg": "FFC7CE", "fg": "9C0006"},
    "Außer Betrieb": {"bg": "FFEB9C", "fg": "9C6500"},
    "Reserve":       {"bg": "D9D9D9", "fg": "404040"},
    "Verloren":      {"bg": "e8c9f5", "fg": "5c007a"},
}

SPALTEN_UEBERSICHT = [
    ("Inventarnummer",     16),
    ("Hersteller",         14),
    ("Modell",             18),
    ("Rufnummer",          16),
    ("SIM-Nummer",         20),
    ("Standort / Ausgabe", 20),
    ("Zustand",            14),
    ("Defektbeschreibung", 35),
    ("Festgestellt am",    16),
    ("Festgestellt von",   20),
    ("Anschaffungsdatum",  18),
    ("Zuletzt geändert",  20),
]

SPALTEN_HISTORIE = [
    ("Datum / Uhrzeit",   20),
    ("Inventarnummer",    16),
    ("Geändertes Feld",   20),
    ("Alter Wert",        25),
    ("Neuer Wert",        25),
    ("Benutzer",          16),
]

FELDER_UEBERSICHT = [
    "inventarnummer", "hersteller", "modell", "rufnummer",
    "sim_nummer", "standort", "zustand", "defekt_beschreibung",
    "defekt_datum", "defekt_gemeldet_von",
    "anschaffungsdatum", "geaendert_am",
]

FELDER_HISTORIE = [
    "geaendert_am", "inventarnummer", "feld",
    "alter_wert", "neuer_wert", "benutzer",
]


def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def export_handys_excel(open_after: bool = True) -> str:
    """
    Exportiert alle Handys + 90-Tage-Historie als Excel-Datei.
    Gibt den Pfad der gespeicherten Datei zurück.
    """
    from functions.handys_db import lade_alle_handys, lade_historie

    handys   = lade_alle_handys()
    historie = lade_historie(tage=90)

    os.makedirs(HANDYS_EXPORT_PATH, exist_ok=True)
    dateiname = f"Handys_Uebersicht_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    pfad = os.path.join(HANDYS_EXPORT_PATH, dateiname)

    wb = openpyxl.Workbook()

    # ── Blatt 1: Übersicht ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Übersicht"

    # Titelzeile (verbundene Zellen A1:J1)
    ws1.merge_cells("A1:J1")
    titel_zelle = ws1["A1"]
    jetzt_str   = datetime.now().strftime("%d.%m.%Y %H:%M")
    titel_zelle.value = (
        f"Diensthandy-Übersicht – DRK Erste-Hilfe-Station Flughafen Köln/Bonn\n"
        f"Erstellt: {jetzt_str} Uhr"
    )
    titel_zelle.font      = Font(name="Segoe UI", bold=True, size=12, color=WEISS)
    titel_zelle.fill      = _header_fill(DRK_ROT)
    titel_zelle.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws1.row_dimensions[1].height = 36

    # Spaltenüberschriften (Zeile 2)
    for col_idx, (ueberschrift, breite) in enumerate(SPALTEN_UEBERSICHT, start=1):
        c = ws1.cell(row=2, column=col_idx, value=ueberschrift)
        c.font      = Font(name="Segoe UI", bold=True, size=10, color=WEISS)
        c.fill      = _header_fill(DRK_DUNKELGRAU)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _thin_border()
        ws1.column_dimensions[get_column_letter(col_idx)].width = breite

    # Datenzeilen (ab Zeile 3)
    for row_idx, handy in enumerate(handys, start=3):
        zeilen_bg = WEISS if row_idx % 2 == 0 else HELLGRAU
        for col_idx, feld in enumerate(FELDER_UEBERSICHT, start=1):
            wert = handy.get(feld, "") or ""
            c = ws1.cell(row=row_idx, column=col_idx, value=wert)
            c.font      = Font(name="Segoe UI", size=10)
            c.alignment = Alignment(vertical="center")
            c.border    = _thin_border()

            # Zustand-Spalte (G = col 7) farbig
            if col_idx == 7:
                farben = ZUSTAND_FARBEN.get(wert, {"bg": WEISS, "fg": DRK_DUNKELGRAU})
                c.fill = _header_fill(farben["bg"])
                c.font = Font(name="Segoe UI", size=10,
                              bold=True, color=farben["fg"])
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.fill = _header_fill(zeilen_bg)

    # Fußzeile (Zusammenfassung)
    fuss_zeile = len(handys) + 4
    zaehler = {z: sum(1 for h in handys if h.get("zustand") == z)
               for z in ["Aktiv", "Defekt", "Außer Betrieb", "Reserve"]}
    ws1.merge_cells(f"A{fuss_zeile}:J{fuss_zeile}")
    fuss = ws1[f"A{fuss_zeile}"]
    fuss.value = (
        f"Gesamt: {len(handys)} Geräte  |  "
        f"Aktiv: {zaehler['Aktiv']}  |  "
        f"Defekt: {zaehler['Defekt']}  |  "
        f"Außer Betrieb: {zaehler['Außer Betrieb']}  |  "
        f"Reserve: {zaehler['Reserve']}"
    )
    fuss.font      = Font(name="Segoe UI", bold=True, size=10, color=WEISS)
    fuss.fill      = _header_fill(DRK_DUNKELGRAU)
    fuss.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[fuss_zeile].height = 22

    # Einfrieren unter Überschrift
    ws1.freeze_panes = "A3"

    # ── Blatt 2: Historie ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Historie")

    # Überschriften
    for col_idx, (ueberschrift, breite) in enumerate(SPALTEN_HISTORIE, start=1):
        c = ws2.cell(row=1, column=col_idx, value=ueberschrift)
        c.font      = Font(name="Segoe UI", bold=True, size=10, color=WEISS)
        c.fill      = _header_fill(DRK_DUNKELGRAU)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _thin_border()
        ws2.column_dimensions[get_column_letter(col_idx)].width = breite

    # Datenzeilen
    for row_idx, eintrag in enumerate(historie, start=2):
        zeilen_bg = WEISS if row_idx % 2 == 0 else HELLGRAU
        for col_idx, feld in enumerate(FELDER_HISTORIE, start=1):
            wert = eintrag.get(feld, "") or ""
            c = ws2.cell(row=row_idx, column=col_idx, value=wert)
            c.font      = Font(name="Segoe UI", size=10)
            c.fill      = _header_fill(zeilen_bg)
            c.alignment = Alignment(vertical="center")
            c.border    = _thin_border()

    ws2.freeze_panes = "A2"

    wb.save(pfad)

    if open_after:
        try:
            os.startfile(pfad)
        except Exception as e:
            print(f"[Handys-Export] Datei konnte nicht geöffnet werden: {e}")

    return pfad

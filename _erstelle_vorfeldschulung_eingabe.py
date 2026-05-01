# -*- coding: utf-8 -*-
"""
Erstellt eine Excel-Eingabetabelle für fehlende Vorfeldschulung-Einträge.
Mitarbeiter ohne Vorfeldschulung-Eintrag werden aufgelistet mit leeren
Feldern für: Datum absolviert | Gültig bis | Bemerkung

Aufruf: python _erstelle_vorfeldschulung_eingabe.py
Ausgabe: Daten/Schulungen/Vorfeldschulung_Eingabe_JJJJMMTT.xlsx  (+ direkt öffnen)
"""
import os
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from functions.schulungen_db import lade_mitarbeiter_mit_schulungen

_BASE = Path(os.path.dirname(os.path.abspath(__file__)))
ZIEL  = _BASE / "Daten" / "Schulungen"
ZIEL.mkdir(parents=True, exist_ok=True)

DATEINAME = ZIEL / f"Vorfeldschulung_Eingabe_{date.today().strftime('%Y%m%d')}.xlsx"

# ── Daten laden ───────────────────────────────────────────────────────────────
alle = lade_mitarbeiter_mit_schulungen()

# Nur Mitarbeiter ohne Vorfeldschulung-Eintrag
fehlende = [
    ma for ma in alle
    if "Vorfeldschulung" not in ma["schulungen"]
]
fehlende.sort(key=lambda m: (m.get("nachname", "").lower(), m.get("vorname", "").lower()))

print(f"Mitarbeiter ohne Vorfeldschulung: {len(fehlende)} von {len(alle)}")

# ── Excel erstellen ────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
ws = wb.active
ws.title = "Vorfeldschulung Eingabe"

# Farben
HEADER_BG    = "1565a8"
HEADER_FG    = "FFFFFF"
EINGABE_BG   = "FFFDE7"   # helles Gelb – Eingabefelder
EINGABE_BG2  = "FFF9C4"   # leicht dunkleres Gelb alternierend
FIXED_BG     = "F5F5F5"   # Grau – gesperrte Felder
FIXED_BG2    = "EBEBEB"

thin   = Side(style="thin", color="AAAAAA")
thick  = Side(style="medium", color="888888")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
border_eingabe = Border(left=thick, right=thick, top=thick, bottom=thick)

SPALTEN = [
    ("Nr.",              6),
    ("Nachname",        20),
    ("Vorname",         16),
    ("Qualifikation",   20),
    ("Datum absolviert", 20),
    ("Gültig bis",       18),
    ("Bemerkung",        30),
]
N_COLS = len(SPALTEN)

# ── Titel ──────────────────────────────────────────────────────────────────────
ws.merge_cells(f"A1:{get_column_letter(N_COLS)}1")
t = ws["A1"]
t.value     = f"Vorfeldschulung – Eintragung fehlender Daten   (Stand: {date.today().strftime('%d.%m.%Y')})"
t.font      = Font(name="Calibri", bold=True, size=14, color=HEADER_FG)
t.fill      = PatternFill("solid", fgColor=HEADER_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# ── Hinweis ────────────────────────────────────────────────────────────────────
ws.merge_cells(f"A2:{get_column_letter(N_COLS)}2")
h = ws["A2"]
h.value     = (
    "Bitte Datum absolviert und Gültig bis eintragen (Format: TT.MM.JJJJ). "
    "Gelbe Felder = Eingabe erforderlich."
)
h.font      = Font(name="Calibri", italic=True, size=10, color="555555")
h.fill      = PatternFill("solid", fgColor="E3F2FD")
h.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 18

# ── Header ─────────────────────────────────────────────────────────────────────
for col, (name, breite) in enumerate(SPALTEN, start=1):
    c = ws.cell(row=3, column=col, value=name)
    c.font      = Font(name="Calibri", bold=True, size=11, color=HEADER_FG)
    c.fill      = PatternFill("solid", fgColor=HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = border
    ws.column_dimensions[get_column_letter(col)].width = breite
ws.row_dimensions[3].height = 22

# ── Datenzeilen ────────────────────────────────────────────────────────────────
for idx, ma in enumerate(fehlende, start=1):
    row = idx + 3
    gerade = (idx % 2 == 0)

    fixed_fill   = PatternFill("solid", fgColor=FIXED_BG2  if gerade else FIXED_BG)
    eingabe_fill = PatternFill("solid", fgColor=EINGABE_BG2 if gerade else EINGABE_BG)

    # Feste Felder: Nr, Nachname, Vorname, Qualifikation
    feste_werte = [
        idx,
        ma.get("nachname", ""),
        ma.get("vorname", ""),
        ma.get("qualifikation", ""),
    ]
    for col, wert in enumerate(feste_werte, start=1):
        c = ws.cell(row=row, column=col, value=wert)
        c.font      = Font(name="Calibri", size=11)
        c.fill      = fixed_fill
        c.border    = border
        c.alignment = Alignment(vertical="center")
        if col == 1:
            c.alignment = Alignment(horizontal="center", vertical="center")

    # Eingabefelder: Datum absolviert (col 5), Gültig bis (col 6), Bemerkung (col 7)
    for col in (5, 6, 7):
        c = ws.cell(row=row, column=col, value="")
        c.font      = Font(name="Calibri", size=11)
        c.fill      = eingabe_fill
        c.border    = border
        c.alignment = Alignment(horizontal="center" if col < 7 else "left", vertical="center")

    ws.row_dimensions[row].height = 18

# ── Summen-Zeile ───────────────────────────────────────────────────────────────
footer_row = len(fehlende) + 4
ws.merge_cells(f"A{footer_row}:{get_column_letter(N_COLS)}{footer_row}")
f = ws.cell(row=footer_row, column=1,
            value=f"Gesamt: {len(fehlende)} Mitarbeiter ohne Vorfeldschulung-Eintrag")
f.font      = Font(name="Calibri", italic=True, size=10, color="555555")
f.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[footer_row].height = 16

# ── Legende ────────────────────────────────────────────────────────────────────
leg_row = footer_row + 2
ws.cell(row=leg_row, column=1, value="Legende:").font = Font(name="Calibri", bold=True, size=10)
leg_items = [
    ("EBEBEB / F5F5F5", "Feste Felder (nicht ändern)"),
    ("FFFDE7 / FFF9C4", "Eingabefelder – bitte ausfüllen"),
]
for i, (farbe, text) in enumerate(leg_items):
    r = leg_row + i + 1
    c1 = ws.cell(row=r, column=1, value=f"  {farbe}")
    c1.font = Font(name="Calibri", size=9, color="888888")
    c2 = ws.cell(row=r, column=2, value=text)
    c2.font = Font(name="Calibri", size=9)

# ── Autofilter + Fixierung ────────────────────────────────────────────────────
ws.auto_filter.ref = f"A3:{get_column_letter(N_COLS)}{len(fehlende) + 3}"
ws.freeze_panes = "A4"

# ── Druckbereich & Seiteneinrichtung ──────────────────────────────────────────
ws.print_area = f"A1:{get_column_letter(N_COLS)}{len(fehlende) + 3}"
ws.page_setup.orientation    = "landscape"
ws.page_setup.paperSize      = 9   # A4
ws.page_setup.fitToPage      = True
ws.page_setup.fitToWidth     = 1
ws.page_setup.fitToHeight    = 0
ws.print_title_rows = "1:3"        # Titel + Hinweis + Header auf jeder Seite

# ── Speichern & öffnen ────────────────────────────────────────────────────────
wb.save(DATEINAME)
print(f"Gespeichert: {DATEINAME}")

import subprocess
subprocess.Popen(["start", "", str(DATEINAME)], shell=True)

"""
Erstellt Inventurliste SanMat als Excel-Datei.
Einmalskript – kein Code geändert.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
import os
from datetime import date


def _set_querformat(ws):
    """Setzt Querformat DIN A4, Fit-to-1-page breit, schmale Ränder."""
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize  = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0   # Höhe beliebig viele Seiten
    ws.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.5, bottom=0.5,
        header=0.2, footer=0.2
    )
    ws.print_options.horizontalCentered = True

# ── Artikel-Daten (identisch zu sanmat_db.py _INITIAL_ARTIKEL) ──────────────
ARTIKEL = [
    ("600400",          "Sofortkältekompresse 15 x 21 cm",                        "Wundversorgung",       "Pck.",  "10 Stück",   ""),
    ("1541 P",          "Steri-Strip 6 x 75 mm",                                  "Wundversorgung",       "Pck.",  "36 Streifen","PZN 03328876"),
    ("9422041",         "Nitril Handschuh Peha-Soft guard XL (lange Stulpe)",      "Schutzausrüstung",     "Pck.",  "100 Stück",  "PZN 03539202"),
    ("9422031",         "Nitril Handschuh Peha-Soft guard L (lange Stulpe)",       "Schutzausrüstung",     "Pck.",  "100 Stück",  "PZN 03539194"),
    ("9422021",         "Nitril Handschuh Peha-Soft guard M (lange Stulpe)",       "Schutzausrüstung",     "Pck.",  "100 Stück",  "PZN 03539188"),
    ("9422011",         "Nitril Handschuh Peha-Soft guard S (lange Stulpe)",       "Schutzausrüstung",     "Pck.",  "100 Stück",  "PZN 03539171"),
    ("920313",          "Alphacheck professional Blutzuckerteststreifen",           "Diagnostik",           "Pck.",  "50 Stück",   "PZN 10329014"),
    ("35408",           "BaSick Bag Spuckbeutel 1500 ml",                          "Patientenversorgung",  "Pck.",  "25 Stück",   "PZN 13818825"),
    ("00-323DS-T060",   "Descosept Sensitive Wipes (20x22 cm)",                    "Desinfektion",         "Pck.",  "60 Tücher",  ""),
    ("00-323DS-OSEB120","Descosept Sensitive Wipes XL Standbeutel (17,5x36 cm)",   "Desinfektion",         "Pck.",  "120 Blatt",  ""),
    ("12562",           "Haft-Fixierbinde 4 cm x 4 m",                            "Wundversorgung",       "Stück", "",           "PZN 04019746"),
    ("12565",           "Haft-Fixierbinde 8 cm x 4 m",                            "Wundversorgung",       "Stück", "",           "PZN 01329067"),
    ("101340",          "Schutzkappen für Thermo Scan 3000/4000/6000",             "Diagnostik",           "Pck.",  "2x20 Stück", "PZN 07437651"),
    ("053030",          "Mulltupfer steril 2x2 Stück (20x20 cm)",                  "Wundversorgung",       "Pck.",  "50 Sets",    "PZN 01364129"),
    ("106600",          "Sterillium Händedesinfektionsmittel 1 Liter",             "Desinfektion",         "Fl.",   "1 Liter",    "PZN 01494079"),
    ("9566",            "Leukosilk Rollenpflaster 1,25 cm x 9,2 m",               "Wundversorgung",       "Rolle", "",           "PZN 04593675"),
    ("9567",            "Leukosilk Rollenpflaster 2,5 cm x 9,2 m",                "Wundversorgung",       "Rolle", "",           "PZN 04593681"),
    ("1222233",         "Laborbecher 0,2 Liter weiß",                              "Verbrauchsmaterial",   "Pck.",  "100 Stück",  ""),
    ("SP-00-S",         "EKG Klebeelektrode Blue Sensor SP-00-S 38mm",             "Diagnostik",           "Pck.",  "50 Stück",   ""),
    ("976802",          "Cutasept F Hautdesinfektion Sprühflasche 250 ml",          "Desinfektion",         "Fl.",   "250 ml",     "PZN 03917271"),
    ("9085",            "Fixomull stretch 10 m x 10 cm",                           "Wundversorgung",       "Stück", "",           "PZN 04539523"),
    ("325012000",       "Ambu SPUR II Beatmungsbeutel Erwachsene",                 "Notfallausrüstung",    "Stück", "",           ""),
    ("719060",          "Lindesa Hautschutzcreme 50 ml",                           "Schutzausrüstung",     "Stück", "50 ml",      "PZN 1281030"),
    ("10670",           "Hansaplast Kinderpflaster Disney Mickey (20 Strips)",      "Wundversorgung",       "Pck.",  "20 Strips",  "PZN 16760150"),
    ("9193006",         "Multi Safe med 6 Kanülensammler ca. 5,1 Liter",           "Entsorgung",           "Stück", "5,1 Liter",  ""),
    ("202210401",       "Infusionstasche S PAX-Light rot (11x25x12 cm)",            "Notfallausrüstung",    "Stück", "",           ""),
    ("22011",           "Spritze Injekt 1 ml 2-teilig ohne Kanüle",                "Verbrauchsmaterial",   "Pck.",  "100 Stück",  "PZN 00896456"),
    ("014211",          "Saugkompresse steril 10x20 cm",                           "Wundversorgung",       "Pck.",  "25 Stück",   "PZN 11606013"),
    ("1003373",         "Aluderm Verbandpäckchen groß 4m x 10cm (Kompr. 10x12cm)", "Wundversorgung",       "Stück", "",           "PZN 03147525"),
    ("01003371",        "Aluderm Verbandpäckchen klein 3m x 6cm (Kompr. 6x8cm)",   "Wundversorgung",       "Stück", "",           ""),
    ("1003372",         "Aluderm Verbandpäckchen mittel 4m x 8cm (Kompr. 8x10cm)", "Wundversorgung",       "Stück", "",           "PZN 03147519"),
    ("10330-2",         "SAM SPLINT-Fingerschiene",                                 "Notfallausrüstung",    "Stück", "",           ""),
    ("10330-1",         "SAM SPLINT Standard 11x92 cm gerollt",                     "Notfallausrüstung",    "Stück", "",           ""),
    ("4063000-100",     "Infusionssystem Intrafix SafeSet 1,8m AirStop",            "Verbrauchsmaterial",   "Kin.",  "100 Stück",  "PZN 01900697"),
    ("973389",          "Bacillol AF 5 Liter Kanister",                             "Desinfektion",         "Kan.",  "5 Liter",    "PZN 00182685"),
    ("1024",            "Leukosilk Rollenpflaster 5 cm x 5 m",                      "Wundversorgung",       "Rolle", "",           "PZN 00626231"),
    ("1022",            "Leukosilk Rollenpflaster 2,5 cm x 5 m",                    "Wundversorgung",       "Rolle", "",           "PZN 00626225"),
    ("672700",          "Bode Messbecher 250 ml (Desinfektionsherstellung)",         "Desinfektion",         "Stück", "250 ml",     "PZN 03650951"),
    ("09160515",        "Alkoholtupfer 6x3 cm gefaltet einzeln verpackt",           "Desinfektion",         "Pck.",  "100 Stück",  "PZN 08468837"),
    ("1003208",         "aluderm Kompresse 10 x 10 cm",                             "Wundversorgung",       "Pck.",  "",           ""),
    ("072585",          "Wundverband steril 7x5 cm",                                "Wundversorgung",       "Pck.",  "50 Stück",   "PZN 07092666"),
    ("40156",           "Wundpflaster 6 cm x 5 m im Spenderkarton",                "Wundversorgung",       "Pck.",  "",           "PZN 04002852"),
    ("1009152",         "aluderm-aluplast Sortiment klein Fingerverband",            "Wundversorgung",       "Pck.",  "",           ""),
    ("1009163",         "aluderm Fingerverband 4x2 cm",                             "Wundversorgung",       "Pck.",  "10 Stück",   ""),
    ("1009184",         "aluderm Fingerkuppenverband 4,3x7,2 cm",                   "Wundversorgung",       "Pck.",  "10 Stück",   ""),
    ("35406",           "Auffangbeutel Universal mit Verschlussring",               "Patientenversorgung",  "Pck.",  "50 Stück",   ""),
    ("70003349",        "Schülke Wipes Wischtücher safe & easy",                    "Desinfektion",         "Karton","6x111 Tücher","PZN 18050464"),
    ("84156",           "Universal Wischtücher blau/weiß 30x33 cm",                "Verbrauchsmaterial",   "Pck.",  "50 Stück",   ""),
    ("16018182",        "Absaugschlauch mit Fingertip und Trichter CH25 (2,0 m)",   "Notfallausrüstung",    "Stück", "",           ""),
    ("0704908210",      "Yankauer Saugansatz CH24 abgewinkelt steril",              "Notfallausrüstung",    "Stück", "",           ""),
    ("1453",            "Cirrus™2 Verneblerset Erwachsene mit EcoLite™ Maske 2,1m", "Notfallausrüstung",   "Stück", "",           ""),
    ("60501502",        "Gänsegurgel gerade 25 cm mit Doppel-Drehkonnektor",        "Notfallausrüstung",    "Stück", "",           ""),
    ("5402822",         "Sauerstoffmaske Erwachsene mit Reservoirbeutel 2m Schlauch","Notfallausrüstung",   "Stück", "",           ""),
    ("000252956",       "Beatmungsmaske Ambu Plus 6 Erwachsene groß",               "Notfallausrüstung",    "Stück", "",           ""),
    ("5402880",         "Hyperventilationsmaske mit Rückatembeutel",                "Notfallausrüstung",    "Stück", "",           ""),
    ("MAD100",          "Medikamentenvernebler MAD 100 Nasalzerstäuber",            "Notfallausrüstung",    "Stück", "",           "PZN 10134233"),
    ("193880",          "Kanülenabwurfbehälter Multi Safe Sani 200 (0,2 Liter)",    "Entsorgung",           "Stück", "0,2 Liter",  ""),
    ("79401805",        "Kombistopfen blau Luer-Lock",                              "Verbrauchsmaterial",   "Pck.",  "100 Stück",  ""),
    ("716270",          "EKG Papier für Corpuls C II (106 mm x 22 m)",              "Diagnostik",           "Stück", "",           ""),
    ("121402",          "Octenisept 500 ml Schleimhautantiseptikum",                "Desinfektion",         "Fl.",   "500 ml",     ""),
    ("972553",          "Baktolan balm Hautschutzcreme 350 ml",                     "Schutzausrüstung",     "Fl.",   "350 ml",     ""),
]

# ── Kategorie-Reihenfolge und Farben ────────────────────────────────────────
KATEGORIEN_INFO = {
    "Wundversorgung":       {"farbe": "FFD6E4BC", "header": "FF4A7C1F"},
    "Desinfektion":         {"farbe": "FFDCE6F1", "header": "FF174069"},
    "Schutzausrüstung":     {"farbe": "FFFFF2CC", "header": "FF7F6000"},
    "Diagnostik":           {"farbe": "FFFCE4D6", "header": "FF843C0C"},
    "Notfallausrüstung":    {"farbe": "FFFFE2CC", "header": "FF833C04"},
    "Patientenversorgung":  {"farbe": "FFE2EFDA", "header": "FF375623"},
    "Verbrauchsmaterial":   {"farbe": "FFEDEDED", "header": "FF595959"},
    "Entsorgung":           {"farbe": "FFFFF0CC", "header": "FF7F5700"},
}

KATEGORIE_REIHENFOLGE = [
    "Wundversorgung",
    "Desinfektion",
    "Schutzausrüstung",
    "Diagnostik",
    "Notfallausrüstung",
    "Patientenversorgung",
    "Verbrauchsmaterial",
    "Entsorgung",
]

# ── Hilfsfunktionen ──────────────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin", color="FFB0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)

def header_border():
    s = Side(style="medium", color="FF666666")
    return Border(left=s, right=s, top=s, bottom=s)

def make_wb():
    wb = openpyxl.Workbook()

    # ── Titelblatt ───────────────────────────────────────────────────────────
    ws_title = wb.active
    ws_title.title = "Deckblatt"
    ws_title.sheet_view.showGridLines = False
    ws_title.column_dimensions["A"].width = 2
    ws_title.column_dimensions["B"].width = 60
    ws_title.column_dimensions["C"].width = 20

    drk_rot = "FFCC0000"
    ws_title.merge_cells("B2:C3")
    c = ws_title["B2"]
    c.value = "DRK – Erste-Hilfe-Station Flughafen Köln/Bonn"
    c.font = Font(name="Calibri", bold=True, size=16, color="FFFFFFFF")
    c.fill = PatternFill("solid", fgColor=drk_rot)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_title.merge_cells("B5:C5")
    c = ws_title["B5"]
    c.value = "Inventurliste Sanitätsmaterial"
    c.font = Font(name="Calibri", bold=True, size=14)
    c.alignment = Alignment(horizontal="center")

    ws_title.merge_cells("B6:C6")
    c = ws_title["B6"]
    c.value = f"Stand: {date.today().strftime('%d.%m.%Y')}"
    c.font = Font(name="Calibri", italic=True, size=11, color="FF555555")
    c.alignment = Alignment(horizontal="center")

    ws_title.merge_cells("B8:C8")
    c = ws_title["B8"]
    c.value = "Inhalt:"
    c.font = Font(name="Calibri", bold=True, size=11)

    for i, kat in enumerate(KATEGORIE_REIHENFOLGE, start=9):
        c = ws_title.cell(row=i, column=2)
        c.value = f"  • {kat}"
        c.font = Font(name="Calibri", size=10)

    ws_title.merge_cells("B18:C18")
    c = ws_title["B18"]
    c.value = "Durchgeführt von: _______________________________"
    c.font = Font(name="Calibri", size=10)

    ws_title.merge_cells("B19:C19")
    c = ws_title["B19"]
    c.value = "Datum: _______________"
    c.font = Font(name="Calibri", size=10)

    ws_title.row_dimensions[2].height = 36
    ws_title.row_dimensions[3].height = 24
    _set_querformat(ws_title)

    # ── Gesamtübersicht ──────────────────────────────────────────────────────
    ws_ges = wb.create_sheet("Gesamtübersicht")
    _build_uebersicht(ws_ges)

    # ── Pro-Kategorie-Blätter ────────────────────────────────────────────────
    for kat in KATEGORIE_REIHENFOLGE:
        ws = wb.create_sheet(kat[:20])   # max 31 Zeichen, safe
        artikel_kat = [a for a in ARTIKEL if a[2] == kat]
        _build_kat_sheet(ws, kat, artikel_kat)

    return wb


def _spaltentitel():
    return [
        "Artikelnr.",        # A
        "Bezeichnung",       # B
        "Einheit",           # C
        "Packungsinhalt",    # D
        "PZN",               # E
        "Min.-Bestand",      # F
        "Soll-Bestand",      # G
        "Ist-Bestand",       # H
        "Lagerort",          # I
        "Ablaufdatum",       # J
        "Bemerkung",         # K
    ]


def _schreibe_header(ws, kat, header_farbe):
    """Kategorieüberschrift + Spaltenköpfe."""
    # Kategorieheader
    ws.merge_cells(f"A1:K1")
    c = ws["A1"]
    c.value = kat
    c.font = Font(name="Calibri", bold=True, size=13, color="FFFFFFFF")
    c.fill = PatternFill("solid", fgColor=header_farbe)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24

    # Spaltenköpfe
    for col, titel in enumerate(_spaltentitel(), start=1):
        c = ws.cell(row=2, column=col, value=titel)
        c.font = Font(name="Calibri", bold=True, size=10)
        c.fill = PatternFill("solid", fgColor="FF404040")
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = header_border()
    ws.row_dimensions[2].height = 30


def _build_kat_sheet(ws, kat, artikel_kat):
    info = KATEGORIEN_INFO.get(kat, {"farbe": "FFEEEEEE", "header": "FF555555"})
    zeilen_farbe = info["farbe"]
    header_farbe = info["header"]

    ws.sheet_view.showGridLines = False

    # Spaltenbreiten
    breiten = [14, 48, 8, 16, 14, 12, 12, 12, 16, 14, 24]
    for col, b in enumerate(breiten, start=1):
        ws.column_dimensions[get_column_letter(col)].width = b

    _schreibe_header(ws, kat, header_farbe)

    row = 3
    alt = False
    for art in sorted(artikel_kat, key=lambda x: x[1]):
        artikelnr, bezeichnung, _, einheit, packinhalt, pzn = art
        # Hintergrundfarbe abwechselnd
        bg = zeilen_farbe if not alt else "FFFFFFFF"
        alt = not alt

        werte = [artikelnr, bezeichnung, einheit, packinhalt, pzn, "", "", "", "", "", ""]
        for col, val in enumerate(werte, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=bg[2:])   # strip 'FF' prefix
            c.font = Font(name="Calibri", size=9)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center", wrap_text=(col == 2))
        ws.row_dimensions[row].height = 16
        row += 1

    # ── Leerzeilen für neue Artikel ─────────────────────────────────────────
    trenn_row = row
    ws.merge_cells(f"A{trenn_row}:K{trenn_row}")
    c = ws.cell(row=trenn_row, column=1,
                value="▼  Neue Artikel im Lager  ▼")
    c.font = Font(name="Calibri", bold=True, size=9, color="FFFFFFFF")
    c.fill = PatternFill("solid", fgColor=header_farbe[2:])
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = header_border()
    ws.row_dimensions[trenn_row].height = 18
    row += 1

    for _ in range(8):
        bg = "FFF5F5F5" if not alt else "FFFFFFFF"
        alt = not alt
        for col in range(1, 12):
            c = ws.cell(row=row, column=col, value="")
            c.fill = PatternFill("solid", fgColor=bg[2:])
            c.border = thin_border()
            c.font = Font(name="Calibri", size=9)
        ws.row_dimensions[row].height = 16
        row += 1

    # Autofilter auf Datenspalten
    ws.auto_filter.ref = f"A2:K{row - 1}"

    # Zeile einfrieren nach Header
    ws.freeze_panes = "A3"

    _set_querformat(ws)


def _build_uebersicht(ws):
    ws.sheet_view.showGridLines = False

    breiten = [2, 26, 48, 8, 16, 14, 12, 12, 12, 16, 14, 24]
    for col, b in enumerate(breiten, start=1):
        ws.column_dimensions[get_column_letter(col)].width = b

    # Hauptheader
    ws.merge_cells("B1:L1")
    c = ws["B1"]
    c.value = f"Inventurliste SanMat – DRK Flughafen Köln/Bonn  |  Stand: {date.today().strftime('%d.%m.%Y')}"
    c.font = Font(name="Calibri", bold=True, size=13, color="FFFFFFFF")
    c.fill = PatternFill("solid", fgColor="FFCC0000")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # Spaltenköpfe
    header_cols = ["Kategorie"] + _spaltentitel()
    for col, titel in enumerate(header_cols, start=2):
        c = ws.cell(row=2, column=col, value=titel)
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor="FF404040")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = header_border()
    ws.row_dimensions[2].height = 30

    row = 3
    for kat in KATEGORIE_REIHENFOLGE:
        info = KATEGORIEN_INFO.get(kat, {"farbe": "FFEEEEEE", "header": "FF555555"})
        artikel_kat = sorted([a for a in ARTIKEL if a[2] == kat], key=lambda x: x[1])

        # Kategorie-Trennzeile
        ws.merge_cells(f"B{row}:L{row}")
        c = ws.cell(row=row, column=2, value=kat)
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=info["header"][2:])
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = header_border()
        ws.row_dimensions[row].height = 18
        row += 1

        alt = False
        for art in artikel_kat:
            artikelnr, bezeichnung, _, einheit, packinhalt, pzn = art
            bg = info["farbe"][2:] if not alt else "FFFFFFFF"
            alt = not alt
            werte = [kat, artikelnr, bezeichnung, einheit, packinhalt, pzn, "", "", "", "", "", ""]
            for col, val in enumerate(werte, start=2):
                c = ws.cell(row=row, column=col, value=val)
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(name="Calibri", size=9)
                c.border = thin_border()
                c.alignment = Alignment(vertical="center", wrap_text=(col == 4))
            ws.row_dimensions[row].height = 15
            row += 1

        # 3 Leerzeilen pro Kategorie für neue Artikel
        for _ in range(3):
            bg = "FFF5F5F5" if not alt else "FFFFFFFF"
            alt = not alt
            for col in range(2, 14):
                c = ws.cell(row=row, column=col, value="")
                c.fill = PatternFill("solid", fgColor=bg)
                c.border = thin_border()
                c.font = Font(name="Calibri", size=9)
            ws.row_dimensions[row].height = 15
            row += 1

    ws.auto_filter.ref = f"B2:L{row - 1}"
    ws.freeze_panes = "B3"

    _set_querformat(ws)


# ── Ausgabe ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    basis = os.path.dirname(os.path.abspath(__file__))
    ausgabe = os.path.join(basis, "Daten", f"Inventurliste_SanMat_{date.today().strftime('%Y-%m-%d')}.xlsx")
    os.makedirs(os.path.dirname(ausgabe), exist_ok=True)

    wb = make_wb()
    wb.save(ausgabe)
    print(f"Gespeichert: {ausgabe}")

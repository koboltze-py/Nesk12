# -*- coding: utf-8 -*-
"""
Erstellt eine Excel-Liste aller aktiven Mitarbeiter, bei denen ein
Vorfeldschulung-Eintrag fehlt (kein Eintrag in schulungseintraege).

Aufruf: python _fehlende_vorfeldschulung.py
Ausgabe: Daten/Schulungen/Fehlende_Vorfeldschulung_JJJJMMTT.xlsx  (+ direkt öffnen)
"""
import os
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from functions.schulungen_db import lade_mitarbeiter_mit_schulungen

_BASE = Path(os.path.dirname(os.path.abspath(__file__)))
ZIEL  = _BASE / "Daten" / "Schulungen"
ZIEL.mkdir(parents=True, exist_ok=True)

DATEINAME = ZIEL / f"Fehlende_Vorfeldschulung_{date.today().strftime('%Y%m%d')}.xlsx"

# ── Daten laden ───────────────────────────────────────────────────────────────
alle = lade_mitarbeiter_mit_schulungen()

# Nur Mitarbeiter ohne Vorfeldschulung-Eintrag
fehlende = [
    ma for ma in alle
    if "Vorfeldschulung" not in ma["schulungen"]
]

# Alphabetisch nach Nachname, Vorname
fehlende.sort(key=lambda m: (m.get("nachname", "").lower(), m.get("vorname", "").lower()))

print(f"Mitarbeiter ohne Vorfeldschulung-Eintrag: {len(fehlende)} von {len(alle)}")

# ── Excel erstellen ────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Fehlende Vorfeldschulung"

HEADER_BG = "1565a8"
HEADER_FG = "FFFFFF"
ROW_ALT   = "EEF4FB"
ROW_NORMAL = "FFFFFF"
FEHLEND_BG = "FFCDD2"   # Rot – Eintrag fehlt

thin   = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

SPALTEN = [
    ("Nr.",           7),
    ("Nachname",     20),
    ("Vorname",      16),
    ("Qualifikation", 20),
]

# Titel-Zeile
ws.merge_cells(f"A1:{get_column_letter(len(SPALTEN))}1")
titel = ws["A1"]
titel.value     = (
    f"Mitarbeiter ohne Vorfeldschulung-Eintrag – Stand: {date.today().strftime('%d.%m.%Y')}"
)
titel.font      = Font(name="Calibri", bold=True, size=14, color=HEADER_FG)
titel.fill      = PatternFill("solid", fgColor=HEADER_BG)
titel.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 26

# Header-Zeile
for col, (name, breite) in enumerate(SPALTEN, start=1):
    c = ws.cell(row=2, column=col, value=name)
    c.font      = Font(name="Calibri", bold=True, size=11, color=HEADER_FG)
    c.fill      = PatternFill("solid", fgColor=HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = border
    ws.column_dimensions[get_column_letter(col)].width = breite
ws.row_dimensions[2].height = 20

# Datenzeilen
fehlend_fill = PatternFill("solid", fgColor=FEHLEND_BG)
for idx, ma in enumerate(fehlende, start=1):
    row = idx + 2
    fill = fehlend_fill if idx % 2 == 1 else PatternFill("solid", fgColor="FFCDD2")
    werte = [
        idx,
        ma.get("nachname", ""),
        ma.get("vorname", ""),
        ma.get("qualifikation", ""),
    ]
    for col, wert in enumerate(werte, start=1):
        c = ws.cell(row=row, column=col, value=wert)
        c.font      = Font(name="Calibri", size=11)
        c.fill      = fehlend_fill
        c.border    = border
        c.alignment = Alignment(vertical="center")
        if col == 1:
            c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 16

# Hinweis-Zeile unten
hinweis_row = len(fehlende) + 3
ws.merge_cells(f"A{hinweis_row}:{get_column_letter(len(SPALTEN))}{hinweis_row}")
h = ws.cell(row=hinweis_row, column=1,
            value=f"Gesamt: {len(fehlende)} Mitarbeiter ohne Vorfeldschulung-Eintrag")
h.font      = Font(name="Calibri", italic=True, size=10, color="555555")
h.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[hinweis_row].height = 16

# Autofilter + Fixierung
ws.auto_filter.ref = f"A2:{get_column_letter(len(SPALTEN))}{len(fehlende) + 2}"
ws.freeze_panes = "A3"

# ── Speichern & öffnen ────────────────────────────────────────────────────────
wb.save(DATEINAME)
print(f"Gespeichert: {DATEINAME}")

import subprocess
subprocess.Popen(["start", "", str(DATEINAME)], shell=True)
